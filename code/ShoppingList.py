#!/usr/bin/env python3
# 
#   ShoppingList.py
#   Reads your .xlsx spreadheet and prints the list to a receipt printer.
##
#   Copyright (C) 2021 by Gregory A. Sanders (dr.gerg@drgerg.com)
#
#   This program is free software: you can redistribute it and/or modify
#   it under the terms of the GNU General Public License as published by
#   the Free Software Foundation, either version 3 of the License, or
#   (at your option) any later version.
#
#   This program is distributed in the hope that it will be useful,
#   but WITHOUT ANY WARRANTY; without even the implied warranty of
#   MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#   GNU General Public License for more details.
#
#   You should have received a copy of the GNU General Public License
#   along with this program.  If not, see <https://www.gnu.org/licenses/>.

import tkinter as tk
from tkinter import Tk,ttk,filedialog,messagebox,simpledialog,INSERT,Toplevel
from tkinter.font import Font
# from tkinter.tix import *
from datetime import datetime
from configparser import ConfigParser
import time,sys,csv,os, textwrap
import openpyxl as xl
import os, warnings, unicodedata, csv, xlrd, pathlib
from os import path
from escpos import *
from escpos.printer import Network
from reportlab.pdfgen.canvas import Canvas


version = "v1.7.2"
confparse = ConfigParser()
path_to_dat = path.abspath(path.join(path.dirname(__file__), 'ShoppingList.ini'))

def main():
    try:
        noteText = getNotes()
        if noteText == "\n":
            text4.delete('1.0','end')
            text4.insert('1.0', 'Notes: ')
        else:
            text4.delete('1.0','end')
            text4.insert('1.0', noteText)
        window.update()
        confparse.read('ShoppingList.ini')
        file = confparse.get('database_loc', 'dbloc')
        ptrIP = confparse.get('printer_address', 'ipaddr')
        pathExists = os.path.exists(file)
        if pathExists == False:
            file = getDataFileLoc()
        if file == "setup":
            file = getDataFileLoc()
            configWindow()
        else:
            text2.delete("1.0", 'end')
            text2.insert("1.0", file + " was selected as your Source database.")
            colName = "Qty"
            final = makeShoppingList(file,colName)
            #
            ## MAKE THE STRING FOR DISPLAY AND PRINTING. SKIP 'NONE' CELLS.
            #
            finalStr = ""
            for tup in final:
                tupString = '(' + str(tup[0]) + ') '
                if tup[1] != None:
                    tupString = tupString + str(tup[1]) + ' '
                if tup[2] != None:
                    tupString = tupString + str(tup[2]) + ' '
                if tup[3] != None:
                    tupString = tupString + str(tup[3]) + "\n"
                else:
                    tupString = tupString + "\n"
                if len(tupString) > 46:
                    tupString = textwrap.fill(tupString, width=46, subsequent_indent='     ') + "\n"
                finalStr = finalStr + tupString
            listFrame = ttk.LabelFrame(window, text="Selected Items:")
            listFrame.grid(column=1, row=2, sticky='ew')
            #
            ##  SHOW THE LIST IN THE LEFT HAND TEXT BOX.
            #
            final_output = tk.Text(listFrame, height = 30, width = 46)
            final_output.grid(column=1, row=2, sticky='ns')
            final_output.insert('1.0',finalStr)
            final_output['state'] = 'disabled'
            text3.delete("1.0", 'end')
            text3.insert(INSERT, "This is your list. If it's empty, add quantities in the spreadsheet & save. Reload, then Print.")
            button_prn = ttk.Button(controlsFrame, text="Print", command=lambda:printIt(finalStr))                  # "Print" button
            button_prn.grid(column=0, row=7, padx=10, pady=10, sticky='n')                                          # Place Print button in grid
            window.update()
    except PermissionError:
        tk.messagebox.showinfo("Database Open Error.", "You must close the database first.")
        window.mainloop()

def getDataFileLoc():
    text2.delete("1.0", 'end')
    text2.insert("1.0", "Navigate to your Source Database file.")
    homeFolder = os.path.expanduser("~")
    confparse.read('ShoppingList.ini')
    file = confparse.get('database_loc', 'dbloc')
    file = filedialog.askopenfilename(initialdir = homeFolder,
                                title = "Navigate to ShoppingList.xlsx Location.",
                                filetypes = (("Excel Files",".xlsx"),))
    confparse.set('database_loc', 'dbloc', file)
    with open('ShoppingList.ini', 'w') as SLcnf:
        confparse.write(SLcnf)
    text2.delete("1.0", 'end')
    text2.insert("1.0", file + " was selected as your Source database. Press 'Reload' to see your list.")
    return file

def getNotes():
    noteText = text4.get('1.0', 'end')
    return noteText

def printIt(final):
    ptrIP = confparse.get('printer_address', 'ipaddr')
    listTitle = confparse.get('list_title', 'text')
    autoPdf = confparse.get('auto_pdf', 'makePDF')
    tnow = datetime.now()
    tnowStr = tnow.strftime("%B %d, %Y %H:%M:%S")
    stuff = getNotes()
    stuffWrap = textwrap.fill(stuff, width=46)
    # print(type(ptrIP))
    # print(ptrIP)
    if ptrIP != '192.168.254.254':
        kitchen = Network(ptrIP)                                #Printer IP Address
        kitchen.set(align='center',width=2,height=2)
        kitchen.text(listTitle + '\n')
        kitchen.set(align='center', width=1,height=1)
        kitchen.text(tnowStr + '\n\n')
        kitchen.set(align='left')
        kitchen.text(stuffWrap)
        kitchen.text("\n\n\n")
        kitchen.text(final)
        kitchen.text("\n\n\n")
        kitchen.cut()
        text2.delete("1.0", 'end')
        text2.insert("1.0", "Your list should be on the printer.\n")
    else:
        text2.delete("1.0", 'end')
        text2.insert("1.0", "The IP address for a receipt printer has not been configured.\n")
    if autoPdf == 'yes':
        pdfPrint(listTitle,tnow,stuffWrap,final)
        text2.insert("2.0", "PDF list was created in your ShoppingList folder.")
    keepGoing = messagebox.askyesno("Hold Up.", "Exit? (Keeps List)")
    if keepGoing == 1:
        exit()

    else:
        text2.delete("1.0", 'end')
        main()

def pdfPrint(lt,tn,stf,fnl):
    listWid = 200                       ## establish a fixed width for the list.
    noteStr = textwrap.wrap(stf, 44)    ## The Note text.
    fnl = fnl.splitlines()
    stfLen = len(noteStr)
    fnlLen = len(fnl)
    # print(fnlLen)
    listLen = (stfLen * 10) + 72 + (fnlLen * 10) + 60
    confparse.read('ShoppingList.ini')
    filePath = confparse.get('database_loc', 'dbloc')
    tnfnStr = tn.strftime("%m%d%y%H%M%S")
    tnowStr = tn.strftime("%B %d, %Y %H:%M:%S")
    filenmStr = '/' + lt + '_' + tnfnStr + '.pdf'
    splitPath = os.path.split(filePath)
    savePath = splitPath[0]
    outputFile = savePath + filenmStr
    # print(outputFile)
    canvas = Canvas(outputFile, pagesize=(listWid,listLen))
    canvas.setFont("Helvetica", 15)                 ## Font for List Title
    tc = (listWid/2)
    prRow = listLen - 40                            ## Start 1 inch (72 pt) below the top
    canvas.drawCentredString(tc,prRow,lt)           ## Print the List Title
    canvas.setFont("Helvetica", 7.5)                ## Font for everthing else
    tnc = (listWid/2)
    prRow = prRow - 10
    canvas.drawCentredString(tnc,prRow,tnowStr)     ## Print the date / time
    prRow = prRow - 24
    for line in noteStr:
        canvas.drawString(10, prRow, line)
        prRow = prRow - 8
    prRow = prRow - 24
    for line in fnl:
        canvas.drawString(10,prRow,line)     ## Print the shopping list data
        prRow = prRow - 8
    canvas.save()



#
## COMPILE THE SHOPPING LIST
#
def makeShoppingList(file,colName):
    # Open the right file
    filename = file
    final = []                                  # list for final output
    ## these next two lines are a sad way to stop getting a
    ## warning about the lack of a default style
    with warnings.catch_warnings(record=True):  
        warnings.simplefilter("always")  
    wb1 = xl.load_workbook(filename, data_only=True)
    for ws1 in wb1.worksheets:
        for c in ws1[1]:                    # Make a list of the column names in this worksheet.
            if c.value == colName:
                cv = c.col_idx              # cv stores column index.
        ##
        ### LOOK THROUGH THE SPREADSHEET AND APPEND EVERY ROW THAT HAS 1 OR MORE IN THE QTY COLUMN.
        ##
        for row in ws1.iter_rows(min_row=2, min_col=cv, max_col=4, values_only=True):    # Check cell validity
            cell = row[0]
            if cell is not None:
                if cell != 0:
                    if cell != " ":
                        if cell >= 1:
                            # print(row)
                            # rowWrappedList = textwrap.wrap(row, width=48)
                            # for row in rowWrappedList:
                            final.append(row)    # Add valid cell contents to final.
    return final



#
## CLEAR SELECTIONS FROM THE SHOPPING LIST SPREADSHEET
#
def clearTheList():
    try:
        confparse.read('ShoppingList.ini')
        file = confparse.get('database_loc', 'dbloc')
        colName = "Qty"
        clearSL = setCtrlVals()[1]
        if clearSL == 1:
            filename = file
            ## these next two lines are a sad way to stop getting a
            ## warning about the lack of a default style
            with warnings.catch_warnings(record=True):  
                warnings.simplefilter("always")  
            wb1 = xl.load_workbook(filename, data_only=True)
            for ws1 in wb1.worksheets:
                thisrow = 2
                for row in ws1.iter_rows(min_row=2, min_col=1, max_col=1):    # Check cell validity
                    ws1.cell(row=thisrow, column=1).value = None
                    thisrow = thisrow + 1
            # saving the destination excel file 
            wb1.save(str(file))
            text2.delete("1.0", 'end')
            text2.insert("1.0", "All quantities are now cleared from the spreadsheet.")
            clearSLVar.set(0)
            tk.messagebox.showinfo("Clear Completed.", "Cleared values from 'Qty'.")
            main()
        else:
            text2.delete("1.0", 'end')
            text2.insert("1.0", "The Clear All Qtys checkbox was not checked, so nothing was changed.")
            keepGoing = tk.messagebox.askyesno("Hold Up.", "Nothing Changed. Go back?")
            if keepGoing == 1:
                text2.delete("1.0", 'end')
                main()
            else:
                exit()
    except PermissionError:
        tk.messagebox.showinfo("Database Open Error.", "You must close the database first.")
        main()
#
## WRAP UP AND DISPLAY
#


def exit():
    sys.exit()

##
##  EVERYTHING SOUTH OF HERE IS THE 'window.mainloop' UNDEFINED BUT YET DEFINED QUASI-FUNCTION
##
window = Tk()  # Create the root window.  'root' is the common name, but I named this one 'window'.
window.title("Shopping List Data Compiler and Printer")  # Set window title
winWd = 1000  # Set window size and placement
winHt = 640
x_Left = int(window.winfo_screenwidth() / 2 - winWd / 2)
y_Top = int(window.winfo_screenheight() / 2 - winHt / 2 - 32)
window.geometry(str(winWd) + "x" + str(winHt) + "+{}+{}".format(x_Left, y_Top))
window.config(background="white")  # Set window background color
window.columnconfigure(0, weight=1)
window.columnconfigure(1, weight=1)
window.columnconfigure(2, weight=1)
window.rowconfigure(0, weight=1)
label_file_explorer = ttk.Label(
    window,  # Create a File Explorer label
    text="Shopping List Data Compiler and Printer",
    width=winWd,
    font=18,
    anchor=tk.CENTER,
    relief='ridge',
    foreground="green",
)
#
##
#
def featureNotReady():
    messagebox.showinfo(title='Not Yet', message='That feature is not ready.')
#
## DEFINE THE CONFIGURE WINDOW
#
def configWindow():
    cw = Toplevel(window)
    cw.title("Configure Options")
    cwinWd = 300  # Set window size and placement
    cwinHt = 440
    x_Left = int(window.winfo_screenwidth() / 2 - cwinWd / 2)
    y_Top = int(window.winfo_screenheight() / 2 - cwinHt / 2)
    cw.config(background="white")  # Set window background color
    cw.geometry(str(cwinWd) + "x" + str(cwinHt) + "+{}+{}".format(x_Left, y_Top))
    cw.columnconfigure(0, weight=1)
    cw.rowconfigure(0, weight=1)
    cw.iconbitmap('./ico/shoppinglist_icon.ico')
    cwlabel = ttk.Label(cw, width=cwinWd, font=18, anchor=tk.CENTER, relief='ridge', text ="Configure Options")
    cwlabel.grid(column=0, row=0, sticky="n")  # Place label in grid

    confparse.read('ShoppingList.ini')
    ptrIP = confparse.get('printer_address', 'ipaddr')
    listTitle = confparse.get('list_title', 'text')
    autoPdf = confparse.get('auto_pdf', 'makePDF')

    def saveConf():
        ptrIP = confIPVar.get()
        listTitle = confTitleVar.get()
        autoPdf = confPDFVar.get()
        autoPdf = autoPdf.lower()
        confparse.set('printer_address', 'ipaddr', ptrIP)
        confparse.set('list_title','text', listTitle)
        confparse.set('auto_pdf','makePDF',autoPdf)
        with open('ShoppingList.ini', 'w') as SLcnf:
            confparse.write(SLcnf)
        text2.delete("1.0", 'end')
        text2.insert("1.0", "Printer IP address " + str(ptrIP) + " was saved in ShoppingList.ini.")
    #
    ## Set up Config text entry boxes
    #
    confIPVar = tk.StringVar()
    confIPLabel = ttk.Label(cw, width=cwinWd, font=('Segoe UI',12), anchor=tk.CENTER, relief='groove', text="Printer IP Address")
    confIPLabel.config(background = 'light blue')
    confIPLabel2 = ttk.Label(cw, width=cwinWd, font=('Segoe UI',10), anchor=tk.CENTER, relief='groove', text="(Use 192.168.254.254 for no printer)")
    confIPLabel2.config(background = 'light yellow')
    confIPEntry = ttk.Entry(cw, justify='center', textvariable = confIPVar, width=18)
    confIPLabel.grid(column=0, row=1, padx=10, pady=10, sticky='n')
    confIPLabel2.grid(column=0, row=2, padx=10, pady=10, sticky='n')
    confIPEntry.grid(column=0, row=3, padx=10, pady=10, sticky='n')
    confIPVar.set(ptrIP)

    confTitleVar = tk.StringVar()
    confTitleLabel = ttk.Label(cw, width=cwinWd, font=('Segoe UI',12), anchor=tk.CENTER, relief='groove', text="List Title")
    confTitleLabel.config(background = 'light blue')
    confTitleEntry = ttk.Entry(cw, justify='center', textvariable = confTitleVar, width=18)
    confTitleLabel.grid(column=0, row=4, padx=10, pady=10, sticky='n')
    confTitleEntry.grid(column=0, row=5, padx=10, pady=10, sticky='n')
    confTitleVar.set(listTitle)

    confPDFVar = tk.StringVar()
    confPDFLabel = ttk.Label(cw, width=cwinWd, font=('Segoe UI',12), anchor=tk.CENTER, relief='groove', text="Auto-create PDF List? (yes/no)")
    confPDFLabel.config(background = 'light blue')
    confPDFEntry = ttk.Entry(cw, justify='center', textvariable = confPDFVar, width=10)
    confPDFLabel.grid(column=0, row=6, padx=10, pady=10, sticky='n')
    confPDFEntry.grid(column=0, row=7, padx=10, pady=10, sticky='n')
    confPDFVar.set(autoPdf)

    cwbutton_cancel = ttk.Button(cw, text="Close", command=cw.destroy)                      # "Close" button
    cwbutton_cancel.grid(column=0, row=8, padx=10, pady=10, sticky='n')                     # Place Close button in grid

    cwbutton_save = ttk.Button(cw, text="Save", command=saveConf)                           # "Save" button
    cwbutton_save.grid(column=0, row=9, padx=10, pady=10, sticky='n')                       # Place Save button in grid
    #

    #


#
## DEFINE THE ABOUT WINDOW
#
def aboutWindow():
    aw = Toplevel(window)
    aw.title("About ShoppingList")
    awinWd = 400  # Set window size and placement
    awinHt = 400
    x_Left = int(window.winfo_screenwidth() / 2 - awinWd / 2)
    y_Top = int(window.winfo_screenheight() / 2 - awinHt / 2)
    aw.config(background="white")  # Set window background color
    aw.geometry(str(awinWd) + "x" + str(awinHt) + "+{}+{}".format(x_Left, y_Top))
    aw.iconbitmap('./ico/shoppinglist_icon.ico')
    awlabel = ttk.Label(aw, width=awinWd, font=18, text ="About " + version)
    awlabel.grid(column=0, columnspan=3, row=0, sticky="n")  # Place label in grid
    aw.columnconfigure(0, weight=1)
    aw.rowconfigure(0, weight=1)
    aboutText = tk.Text(aw, height=20, width=170, bd=3, padx=10, pady=10, wrap='word', font=nnFont)
    aboutText.grid(column=0, row=1)
    aboutText.insert(INSERT, "ShoppingList reads your specified LibreOffice Calc or Excel spreadsheet and compiles your shopping list for printing on a receipt printer.\n\n"
    "The project can be found on GitHub at https://github.com/casspop/shoppingList"
    "\n\n- Greg Sanders\n\nThis app is written in Python and compiled using PyInstaller.\n\n"
"Check out more of my projects at www.drgerg.com.")
#
## DEFINE THE HELP WINDOW
#
def helpWindow():

    hw = Toplevel(window)
    hw.title("Help")
    hwinWd = 400  # Set window size and placement
    hwinHt = 600
    x_Left = int(window.winfo_screenwidth() / 2 - hwinWd / 2)
    y_Top = int(window.winfo_screenheight() / 2 - hwinHt / 2)
    hw.config(background="white")  # Set window background color
    hw.geometry(str(hwinWd) + "x" + str(hwinHt) + "+{}+{}".format(x_Left, y_Top))
    hw.iconbitmap('./ico/shoppinglist_icon.ico')
    hwlabel = ttk.Label(hw, width=hwinWd, font=18, text ="Help")
    hw.columnconfigure(0, weight=1)
    hw.rowconfigure(0, weight=1)
    hw.rowconfigure(1, weight=1)

    helpText = tk.Text(hw, height=40, width=80, bd=3, padx=6, pady=6, wrap='word', font=nnFont)
    helpsb = ttk.Scrollbar(hw, orient='vertical', command=helpText.yview)

    helpText['yscrollcommand'] = helpsb.set
    hwlabel.grid(column=0, columnspan=3, row=0, padx=10, pady=10)  # Place label in grid
    helpText.grid(column=0, row=1)
    helpsb.grid(column=1, row=1, sticky='ns')
    with open("helpText.txt", "r") as f:
        hlptxt = f.read()
    helpText.insert(INSERT, hlptxt)
#
## Open the data file for editing
#
def editData():
    confparse.read('ShoppingList.ini')
    file = confparse.get('database_loc', 'dbloc')
    try:
        os.startfile(file)
    except PermissionError:
        tk.messagebox.showinfo("Database Open Error.", "You must close the database first.")


#
# "Options" frames them nicely.
# #
def setCtrlVals():
    colName = "Qty"
    clearSL = clearSLVar.get()
    return colName, clearSL

#
## MENU AND MENU ITEMS
#
tk.Frame(window)
menu = tk.Menu(window)
window.config(menu=menu)
nnFont = Font(family="Segoe UI", size=10)          ## Set the base font
fileMenu = tk.Menu(menu, tearoff=False)
fileMenu.add_command(label="Configure", command=lambda:configWindow())
fileMenu.add_command(label="Select Database", command=getDataFileLoc)
fileMenu.add_command(label="Exit", command=exit)
menu.add_cascade(label="File", menu=fileMenu)

editMenu = tk.Menu(menu, tearoff=False)
editMenu.add_command(label="Help", command=helpWindow)
editMenu.add_command(label="About", command=aboutWindow)
menu.add_cascade(label="Help", menu=editMenu)
#
## VARIOUS ATTEMPTS AT GETTING WINDOWS ICONS TO WORK
#
# The secret was getting the path stuff right in the pyinstaller command line: pyinstaller --add-binary=".\ico\shoppinglist_icon.ico;ico" --noconsole --icon=shoppinglist_icon.ico NextNum.py
#
window.iconbitmap('./ico/shoppinglist_icon.ico')

controlsFrame = ttk.LabelFrame(window, text="Options")             # larger frame to hold Radio Button frame
controlsFrame.grid(column=0, row=2, padx=20, sticky='nw')
#
## Set up push-buttons
#
button_clear = ttk.Button(controlsFrame, text="Clear All", command=clearTheList)        # "Clear" button
button_clear.grid(column=0, row=9, padx=10, pady=10, sticky='n')                        # Place Clear button in grid
button_reload = ttk.Button(controlsFrame, text="Reload", command=main)                  # "Reload" button
button_reload.grid(column=0, row=8, padx=10, pady=10, sticky='n')                       # Place Reload button in grid
button_editData = ttk.Button(controlsFrame, text="Edit Database", command=editData)                  # "Reload" button
button_editData.grid(column=0, row=10, padx=10, pady=10, sticky='n')                       # Place Reload button in grid
button_exit = ttk.Button(controlsFrame, text="Exit", command=exit)                      # "Exit" button
button_exit.grid(column=0, row=11, padx=10, pady=10, sticky='n')                        # Place Exit button in grid
#
# editDataTip = Balloon(window)
# editDataTip.bind_widget(button_editData,balloonmsg="Open your database, make your edit, close database.")
#
## Set up check boxes
#
clearSLVar = tk.IntVar(value=0)
clearSLChkBox = tk.Checkbutton(controlsFrame,text='Clear All Qtys.', variable=clearSLVar, onvalue=1, offvalue=0, command=setCtrlVals)      # define it
clearSLChkBox.grid(column=0, row=4, sticky='nw')                                                   # place it
#
## Set up Notes and Reminders box
#

T4Frame = ttk.LabelFrame(window, text="Notes and Reminders:")             # Frame for list Notes.
T4Frame.grid(column=2, row=2, sticky='ew')
#
## Set up text windows
#
# text1 = tk.Text(window, height=6, width=150, wrap='word', font=nnFont)
text2 = tk.Text(window, height=2, width=150, font=nnFont)
text3 = tk.Text(window, height=3, width=150, font=nnFont)
text4 = tk.Text(T4Frame, height=28, width=46, wrap='word', font=nnFont)

label_file_explorer.grid(column=0, columnspan=7, row=0, sticky="n")  # Place label in grid

# text1.grid(column=0, columnspan=7, row=1, padx=10)
text2.grid(column=0, columnspan=7, row=6, padx=10)
text3.grid(column=0, columnspan=7, row=7, padx=10)
text4.grid(column=1, row=1)
main()
window.mainloop()  # Run the (not defined with 'def') main window loop.
